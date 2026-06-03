"""
test_merge_new_excel.py
=======================
merge_new_excel.py 的整合測試（black-box integration test）。

策略：
  - 在臨時資料夾中放置一份 merge_new_excel.py 的複本與測試用 xlsx 檔案，
  - 以子行程執行該腳本（腳本以自身 __file__ 所在資料夾為工作目錄），
  - 驗證產生的 excel001-v001.xlsx 內容是否正確合併。

執行方式：
    python -m unittest test_merge_new_excel -v
"""

import os
import shutil
import subprocess
import sys
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SCRIPT = os.path.join(HERE, "merge_new_excel.py")


def _write_single_col(path, header, values):
    """寫一個只有 A 欄的 xlsx：第 1 列為 header，其後為 values。"""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=header)
    for i, v in enumerate(values, start=2):
        ws.cell(row=i, column=1, value=v)
    wb.save(path)


def _write_base(path, headers):
    """寫基底檔，第 1 列為多個欄位標題。"""
    wb = Workbook()
    ws = wb.active
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    wb.save(path)


def _run_script(folder):
    """在 folder 中執行腳本複本；以 stdin 餵入 Enter 以略過 input() 暫停。"""
    script_copy = os.path.join(folder, "merge_new_excel.py")
    shutil.copyfile(SCRIPT, script_copy)
    proc = subprocess.run(
        [sys.executable, script_copy],
        input="\n",
        capture_output=True,
        text=True,
        timeout=120,
    )
    return proc


class MergeNewExcelTest(unittest.TestCase):
    def setUp(self):
        self.folder = tempfile.mkdtemp(prefix="merge_test_")
        self.addCleanup(shutil.rmtree, self.folder, ignore_errors=True)

    def _path(self, name):
        return os.path.join(self.folder, name)

    def test_merges_new_files_as_columns_in_serial_order(self):
        # 基底：3 個既有欄位
        _write_base(self._path("excel001.xlsx"), ["頁碼", "段落編號", "內容"])
        # 兩個新檔，刻意以非排序順序建立，但檔名流水號決定合併順序
        _write_single_col(self._path("excel#2.xlsx"), "h2", ["b1", "b2"])
        _write_single_col(self._path("excel#1.xlsx"), "h1", ["a1", "a2", "a3"])

        proc = _run_script(self.folder)
        self.assertEqual(proc.returncode, 0, msg=proc.stdout + proc.stderr)

        out_path = self._path("excel001-v001.xlsx")
        self.assertTrue(os.path.exists(out_path), "未產生 excel001-v001.xlsx")

        ws = load_workbook(out_path).active

        # 原 3 欄保留 + 2 個新欄 = 5 欄
        self.assertEqual(ws.max_column, 5)
        self.assertEqual(ws.cell(row=1, column=1).value, "頁碼")
        self.assertEqual(ws.cell(row=1, column=2).value, "段落編號")
        self.assertEqual(ws.cell(row=1, column=3).value, "內容")

        # 第 4 欄應為 excel#1（流水號較小者先合併），第 5 欄為 excel#2
        self.assertTrue(ws.cell(row=1, column=4).value.startswith("excel#1_"))
        self.assertTrue(ws.cell(row=1, column=5).value.startswith("excel#2_"))

        # 來源 A 欄（含其自身標題列）會從輸出第 2 列開始填入
        col4 = [ws.cell(row=r, column=4).value for r in range(2, 6)]
        self.assertEqual(col4, ["h1", "a1", "a2", "a3"])
        col5 = [ws.cell(row=r, column=5).value for r in range(2, 5)]
        self.assertEqual(col5, ["h2", "b1", "b2"])

    def test_skips_already_merged_files(self):
        # 基底已含名為 excel#1 的合併欄位 → 該檔應被略過
        _write_base(
            self._path("excel001.xlsx"),
            ["頁碼", "段落編號", "內容", "excel#1_202601010000"],
        )
        _write_single_col(self._path("excel#1.xlsx"), "h1", ["a1"])

        proc = _run_script(self.folder)
        self.assertEqual(proc.returncode, 0, msg=proc.stdout + proc.stderr)

        # 沒有新檔需合併 → 不應產生新版本檔
        self.assertFalse(os.path.exists(self._path("excel001-v001.xlsx")))
        self.assertIn("無新檔案需要合併", proc.stdout)

    def test_uses_latest_version_as_base(self):
        # 已存在 v001、v002 → 應以 v002 為基底，產生 v003
        _write_base(self._path("excel001.xlsx"), ["頁碼", "段落編號", "內容"])
        _write_base(self._path("excel001-v001.xlsx"), ["頁碼", "段落編號", "內容"])
        _write_base(
            self._path("excel001-v002.xlsx"),
            ["頁碼", "段落編號", "內容", "old_202601010000"],
        )
        _write_single_col(self._path("excel#5.xlsx"), "h5", ["x1"])

        proc = _run_script(self.folder)
        self.assertEqual(proc.returncode, 0, msg=proc.stdout + proc.stderr)

        out_path = self._path("excel001-v003.xlsx")
        self.assertTrue(os.path.exists(out_path), "未以最新版本為基底產生 v003")

        ws = load_workbook(out_path).active
        # 繼承 v002 的 4 欄 + 1 新欄
        self.assertEqual(ws.max_column, 5)
        self.assertEqual(ws.cell(row=1, column=4).value, "old_202601010000")
        self.assertTrue(ws.cell(row=1, column=5).value.startswith("excel#5_"))


if __name__ == "__main__":
    unittest.main(verbosity=2)
