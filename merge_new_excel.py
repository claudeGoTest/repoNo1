"""
merge_new_excel.py
==================
自動把資料夾中所有新增的 excel*.xlsx 檔案，依檔名流水號順序，
依次新增欄位合併到最新版本檔，並另存為新版號 excel001-v(NNN+1).xlsx。

需求：Python 3.7+；openpyxl（缺則自動安裝）
"""

import os
import re
import sys
import glob
from datetime import datetime, timezone, timedelta


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass
    print("[提示] 未安裝 openpyxl 套件，正在自動安裝...")
    import subprocess
    cmds = [
        [sys.executable, "-m", "pip", "install", "--user", "openpyxl"],
        [sys.executable, "-m", "pip", "install", "openpyxl"],
    ]
    for cmd in cmds:
        print("  > " + " ".join(cmd))
        rc = subprocess.call(cmd)
        if rc == 0:
            break
    try:
        import openpyxl  # noqa: F401
        print("[OK] openpyxl 安裝完成。\n")
    except ImportError:
        print("[錯誤] openpyxl 自動安裝失敗，請手動執行：")
        print("  " + sys.executable + " -m pip install openpyxl")
        try:
            input("\n按 Enter 鍵結束...")
        except Exception:
            pass
        sys.exit(1)


_ensure_openpyxl()

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def main():
    folder = os.path.dirname(os.path.abspath(__file__))
    print("工作資料夾：" + folder)

    pattern = re.compile(r"^excel001-v(\d+)$", re.IGNORECASE)
    max_v, base_path = 0, None
    for p in glob.glob(os.path.join(folder, "excel001-v*.xlsx")):
        name = os.path.splitext(os.path.basename(p))[0]
        m = pattern.match(name)
        if m:
            v = int(m.group(1))
            if v > max_v:
                max_v, base_path = v, p

    if base_path is None:
        base_path = os.path.join(folder, "excel001.xlsx")
        if not os.path.exists(base_path):
            print("[錯誤] 找不到基底檔案 excel001.xlsx 或任何 excel001-vNNN.xlsx")
            input("按 Enter 鍵結束...")
            sys.exit(1)

    new_v = max_v + 1
    new_name = "excel001-v{:03d}.xlsx".format(new_v)
    new_path = os.path.join(folder, new_name)
    print("基底：" + os.path.basename(base_path))
    print("目標：" + new_name)

    wb = load_workbook(base_path)
    ws = wb.active
    existing = set()
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            existing.add(str(v).split("_")[0])
    merged_list = sorted(existing - {"頁碼", "段落編號", "內容"})
    print("已合併過：" + str(merged_list))

    candidates = []
    for p in glob.glob(os.path.join(folder, "excel*.xlsx")):
        name = os.path.splitext(os.path.basename(p))[0]
        if name == "excel001":
            continue
        if pattern.match(name):
            continue
        if name.startswith("~$"):
            continue
        if name in existing:
            continue
        m = re.search(r"(\d+)", name)
        serial = int(m.group(1)) if m else 0
        candidates.append((serial, name, p))

    candidates.sort(key=lambda x: x[0])

    if not candidates:
        print("\n無新檔案需要合併。")
        try:
            input("\n按 Enter 鍵結束...")
        except Exception:
            pass
        return

    print("\n發現 " + str(len(candidates)) + " 個新檔案：" + str([n for _, n, _ in candidates]))

    tw = timezone(timedelta(hours=8))
    ts = datetime.now(tw).strftime("%Y%m%d%H%M")
    print("時間戳：" + ts)

    header_font = Font(bold=True, name="DengXian", size=12)
    content_font = Font(name="DengXian", size=11)
    added = []

    for serial, name, path in candidates:
        new_col = ws.max_column + 1
        wb_n = load_workbook(path)
        ws_n = wb_n.active

        header_text = name + "_" + ts
        hdr = ws.cell(row=1, column=new_col, value=header_text)
        hdr.font = header_font
        hdr.alignment = Alignment(wrap_text=True, vertical="top")

        for r_idx, row in enumerate(ws_n.iter_rows(min_row=1, max_col=1, values_only=True), start=2):
            c = ws.cell(row=r_idx, column=new_col, value=row[0])
            c.font = content_font
            c.alignment = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions[get_column_letter(new_col)].width = 20
        added.append(header_text)
        print("  + 欄 " + str(new_col) + ": " + header_text + "  (" + str(ws_n.max_row) + " 筆)")

    wb.save(new_path)
    print("\n已建立：" + new_path)
    print("最終尺寸：" + str(ws.max_row) + " 列 x " + str(ws.max_column) + " 欄")

    print("\n新增欄位清單：")
    for h in added:
        print("  - " + h)

    try:
        input("\n按 Enter 鍵結束...")
    except Exception:
        pass


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("\n[錯誤] " + type(e).__name__ + ": " + str(e))
        import traceback
        traceback.print_exc()
        try:
            input("\n按 Enter 鍵結束...")
        except Exception:
            pass
        sys.exit(1)
